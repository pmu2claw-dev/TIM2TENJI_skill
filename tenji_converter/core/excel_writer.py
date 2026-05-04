from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook


TEST_NOTE_HEADERS = {
    "B": "No.",
    "C": "Category",
    "D": "Scenario",
    "E": "Power",
    "F": "Block",
    "G": "Symbol",
    "H": "Measure_Condition",
    "I": "Expected",
    "J": "Lower",
    "K": "Upper",
    "L": "Unit",
    "M": "Judge",
    "N": "Comment",
}


def ensure_workbook(path: Path, template: Path | None = None):
    if path.exists():
        return load_workbook(path, keep_vba=True)
    if template and template.exists():
        return load_workbook(template, keep_vba=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Test_Note"
    for col, header in TEST_NOTE_HEADERS.items():
        ws[f"{col}1"] = header
    wb.create_sheet("Revision")
    return wb


def commands_to_measure_cond(commands: list[dict[str, Any]]) -> str:
    lines: list[str] = []
    for cmd in commands:
        name = str(cmd.get("cmd", "")).strip()
        value = str(cmd.get("value", "")).strip()
        if value:
            lines.append(f"{name} {value}")
        else:
            lines.append(name)
    return "\n".join([x for x in lines if x])


def _append_revision(ws_rev, message: str):
    row = ws_rev.max_row + 1 if ws_rev.max_row >= 1 else 1
    ws_rev[f"A{row}"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws_rev[f"B{row}"] = message


def write_spec_to_excel(
    spec: dict[str, Any],
    output_path: Path,
    template: Path | None = None,
    existing_excel: Path | None = None,
) -> Path:
    action = spec.get("action", "generate")
    target = existing_excel if action == "append" and existing_excel else output_path

    wb = ensure_workbook(target if target else output_path, template=template)
    ws = wb["Test_Note"] if "Test_Note" in wb.sheetnames else wb.active

    if ws.max_row < 1:
        for col, header in TEST_NOTE_HEADERS.items():
            ws[f"{col}1"] = header

    for item in spec.get("items", []):
        row = ws.max_row + 1
        ws[f"B{row}"] = row - 1
        ws[f"C{row}"] = item.get("category", "General")
        ws[f"D{row}"] = item.get("scenario", "")
        ws[f"E{row}"] = item.get("power", "")
        ws[f"F{row}"] = item.get("block", "")
        ws[f"G{row}"] = item.get("symbol", "")
        ws[f"H{row}"] = item.get("measure_cond") or commands_to_measure_cond(item.get("commands", []))
        ws[f"I{row}"] = item.get("expected", "")
        ws[f"J{row}"] = item.get("lower", "")
        ws[f"K{row}"] = item.get("upper", "")
        ws[f"L{row}"] = item.get("unit", "")
        ws[f"M{row}"] = item.get("judge", "")
        ws[f"N{row}"] = item.get("comment", "")

    if "Revision" not in wb.sheetnames:
        wb.create_sheet("Revision")
    _append_revision(wb["Revision"], f"{action}: {len(spec.get('items', []))} item(s)")

    save_path = output_path if action == "generate" else (existing_excel or output_path)
    save_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(save_path)
    return save_path
